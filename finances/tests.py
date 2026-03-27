import csv
import io
from datetime import date
from decimal import Decimal

import pytest
from django.contrib.auth.models import User
from django.urls import reverse

from finances.models import Account, Category, Transaction
from finances.services import _auto_categorize, process_bank_statement_import


class TestAccountModel:
    """Testes para o modelo Account."""

    def test_account_str(self, db, usuario):
        conta = Account.objects.create(
            nome='Nubank', tipo='conta_corrente', usuario=usuario
        )
        assert str(conta) == 'Nubank (Conta Corrente)'

    def test_account_tipos_validos(self, db, usuario):
        for tipo in ['conta_corrente', 'carteira', 'cartao_credito']:
            conta = Account.objects.create(
                nome=f'Conta {tipo}', tipo=tipo, usuario=usuario
            )
            assert conta.tipo == tipo

    def test_account_unique_per_user(self, db, usuario):
        from django.db import IntegrityError
        Account.objects.create(nome='Carteira', tipo='carteira', usuario=usuario)
        with pytest.raises(IntegrityError):
            Account.objects.create(nome='Carteira', tipo='carteira', usuario=usuario)


class TestCategoryKeywords:
    """Testa o método get_keywords_list de Category."""

    def test_keywords_list_vazia(self, db, usuario):
        cat = Category.objects.create(nome='Sem KW', tipo='saida', usuario=usuario)
        assert cat.get_keywords_list() == []

    def test_keywords_list_simples(self, db, usuario):
        cat = Category.objects.create(
            nome='Transporte', tipo='saida', usuario=usuario, keywords='uber,99,taxi'
        )
        assert 'uber' in cat.get_keywords_list()
        assert 'taxi' in cat.get_keywords_list()

    def test_keywords_list_case_insensitive(self, db, usuario):
        cat = Category.objects.create(
            nome='Alimentação', tipo='saida', usuario=usuario, keywords='McDonald,BURGER KING'
        )
        assert 'mcdonald' in cat.get_keywords_list()
        assert 'burger king' in cat.get_keywords_list()


class TestAutoCategorize:
    """Testa a função _auto_categorize do services.py."""

    def test_categoriza_por_keyword(self, db, usuario):
        cat = Category.objects.create(
            nome='Transporte', tipo='saida', usuario=usuario, keywords='uber,taxi'
        )
        resultado = _auto_categorize('UBER TRIP SP', usuario)
        assert resultado.pk == cat.pk

    def test_fallback_avulsa(self, db, usuario):
        resultado = _auto_categorize('TARIFA BANCARIA', usuario)
        assert resultado.nome == 'Avulsa'

    def test_avulsa_criada_automaticamente(self, db, usuario):
        assert not Category.objects.filter(nome='Avulsa', usuario=usuario).exists()
        _auto_categorize('DEBITO AVULSO', usuario)
        assert Category.objects.filter(nome='Avulsa', usuario=usuario).exists()

    def test_avulsa_nao_duplicada(self, db, usuario):
        _auto_categorize('TX 1', usuario)
        _auto_categorize('TX 2', usuario)
        assert Category.objects.filter(nome='Avulsa', usuario=usuario).count() == 1

    def test_keyword_case_insensitive(self, db, usuario):
        cat = Category.objects.create(
            nome='Alimentação', tipo='saida', usuario=usuario, keywords='ifood,rappi'
        )
        resultado = _auto_categorize('IFOOD PEDIDO 12345', usuario)
        assert resultado.pk == cat.pk


class TestProcessBankStatementCSV:
    """Testa a importação via CSV."""

    def _make_csv_file(self, rows: list, filename='extrato.csv'):
        """Cria objeto de arquivo CSV simulado."""
        fieldnames = ['data', 'descricao', 'valor']
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        content = buf.getvalue().encode('utf-8')

        class FakeFile:
            name = filename

            def read(self):
                return content

        return FakeFile()

    def test_csv_cria_transacoes(self, db, usuario):
        conta = Account.objects.create(nome='Banco', tipo='conta_corrente', usuario=usuario)
        arquivo = self._make_csv_file([
            {'data': '2026-03-01', 'descricao': 'Salário', 'valor': '3000.00'},
            {'data': '2026-03-05', 'descricao': 'Mercado', 'valor': '-500.00'},
        ])
        resultado = process_bank_statement_import(arquivo, usuario, conta)
        assert resultado['criadas'] == 2
        assert resultado['ignoradas'] == 0

    def test_csv_auto_categoriza(self, db, usuario):
        Category.objects.create(
            nome='Supermercado', tipo='saida', usuario=usuario, keywords='mercado,supermercado'
        )
        conta = Account.objects.create(nome='Banco', tipo='conta_corrente', usuario=usuario)
        arquivo = self._make_csv_file([
            {'data': '2026-03-10', 'descricao': 'SUPERMERCADO XYZ', 'valor': '-200.00'},
        ])
        process_bank_statement_import(arquivo, usuario, conta)
        t = Transaction.objects.filter(usuario=usuario).first()
        assert t.categoria.nome == 'Supermercado'

    def test_csv_fallback_avulsa(self, db, usuario):
        conta = Account.objects.create(nome='Banco', tipo='conta_corrente', usuario=usuario)
        arquivo = self._make_csv_file([
            {'data': '2026-03-15', 'descricao': 'TARIFA BANCARIA', 'valor': '-10.00'},
        ])
        process_bank_statement_import(arquivo, usuario, conta)
        t = Transaction.objects.filter(usuario=usuario).first()
        assert t.categoria.nome == 'Avulsa'

    def test_formato_nao_suportado(self, db, usuario):
        conta = Account.objects.create(nome='Banco', tipo='conta_corrente', usuario=usuario)

        class FakeFile:
            name = 'extrato.xlsx'

            def read(self):
                return b'nada'

        resultado = process_bank_statement_import(FakeFile(), usuario, conta)
        assert resultado['criadas'] == 0
        assert len(resultado['erros']) > 0


# ============================================================
# Fase 3 — Services e Selectors
# ============================================================

class TestCreateTransaction:
    """Testes unitários para create_transaction (independente de HTTP)."""

    def test_cria_transacao_valida(self, db, usuario):
        from finances.services import create_transaction
        data = {
            'valor': '500.00',
            'data': date.today(),
            'tipo': 'entrada',
            'descricao': 'Salário',
        }
        t = create_transaction(usuario, data)
        assert t.pk is not None
        assert t.valor == Decimal('500.00')
        assert t.tipo == 'entrada'
        assert t.usuario == usuario

    def test_cria_transacao_valor_zero_levanta_erro(self, db, usuario):
        from django.core.exceptions import ValidationError
        from finances.services import create_transaction
        with pytest.raises(ValidationError, match='valor'):
            create_transaction(usuario, {
                'valor': '0.00',
                'data': date.today(),
                'tipo': 'saida',
                'descricao': 'Inválido',
            })

    def test_cria_transacao_valor_negativo_levanta_erro(self, db, usuario):
        from django.core.exceptions import ValidationError
        from finances.services import create_transaction
        with pytest.raises(ValidationError):
            create_transaction(usuario, {
                'valor': '-100.00',
                'data': date.today(),
                'tipo': 'saida',
                'descricao': 'Inválido',
            })

    def test_cria_transacao_tipo_invalido_levanta_erro(self, db, usuario):
        from django.core.exceptions import ValidationError
        from finances.services import create_transaction
        with pytest.raises(ValidationError, match='tipo'):
            create_transaction(usuario, {
                'valor': '100.00',
                'data': date.today(),
                'tipo': 'receita',  # tipo inválido
                'descricao': 'Inválido',
            })

    def test_cria_transacao_com_categoria(self, db, usuario):
        from finances.services import create_transaction
        cat = Category.objects.create(nome='Salário', tipo='entrada', usuario=usuario)
        t = create_transaction(usuario, {
            'valor': '3000.00',
            'data': date.today(),
            'tipo': 'entrada',
            'descricao': 'CLT',
            'categoria_id': cat.pk,
        })
        assert t.categoria == cat

    def test_cria_transacao_categoria_outro_usuario_levanta_erro(self, db, usuario):
        from django.core.exceptions import ValidationError
        from finances.services import create_transaction
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro', password='pass123')
        cat = Category.objects.create(nome='Cat Outro', tipo='saida', usuario=outro)
        with pytest.raises(ValidationError):
            create_transaction(usuario, {
                'valor': '100.00',
                'data': date.today(),
                'tipo': 'saida',
                'descricao': 'Tentativa indevida',
                'categoria_id': cat.pk,
            })


class TestUpdateTransaction:
    """Testes unitários para update_transaction."""

    def _criar_transacao(self, usuario):
        from finances.services import create_transaction
        return create_transaction(usuario, {
            'valor': '200.00',
            'data': date.today(),
            'tipo': 'saida',
            'descricao': 'Original',
        })

    def test_atualiza_descricao(self, db, usuario):
        from finances.services import update_transaction
        t = self._criar_transacao(usuario)
        atualizada = update_transaction(t.pk, usuario, {'descricao': 'Nova descrição'})
        assert atualizada.descricao == 'Nova descrição'

    def test_atualiza_valor(self, db, usuario):
        from finances.services import update_transaction
        t = self._criar_transacao(usuario)
        atualizada = update_transaction(t.pk, usuario, {'valor': '999.99'})
        assert atualizada.valor == Decimal('999.99')

    def test_update_outro_usuario_levanta_erro(self, db, usuario):
        from finances.services import update_transaction
        from django.contrib.auth.models import User
        from finances.models import Transaction
        t = self._criar_transacao(usuario)
        outro = User.objects.create_user(username='outro2', password='pass123')
        with pytest.raises(Transaction.DoesNotExist):
            update_transaction(t.pk, outro, {'descricao': 'Hack'})


class TestDeleteTransaction:
    """Testes unitários para delete_transaction."""

    def test_deleta_transacao(self, db, usuario):
        from finances.services import create_transaction, delete_transaction
        from finances.models import Transaction
        t = create_transaction(usuario, {
            'valor': '50.00',
            'data': date.today(),
            'tipo': 'saida',
            'descricao': 'Para deletar',
        })
        pk = t.pk
        delete_transaction(pk, usuario)
        assert not Transaction.objects.filter(pk=pk).exists()

    def test_delete_outro_usuario_levanta_erro(self, db, usuario):
        from finances.services import create_transaction, delete_transaction
        from finances.models import Transaction
        from django.contrib.auth.models import User
        t = create_transaction(usuario, {
            'valor': '50.00',
            'data': date.today(),
            'tipo': 'saida',
            'descricao': 'Para deletar',
        })
        outro = User.objects.create_user(username='outro3', password='pass123')
        with pytest.raises(Transaction.DoesNotExist):
            delete_transaction(t.pk, outro)


class TestCalculateBalance:
    """Testes unitários para calculate_balance."""

    def test_saldo_zerado_sem_transacoes(self, db, usuario):
        from finances.services import calculate_balance
        saldo = calculate_balance(usuario)
        assert saldo['total_entradas'] == Decimal('0.00')
        assert saldo['total_saidas'] == Decimal('0.00')
        assert saldo['saldo_liquido'] == Decimal('0.00')

    def test_saldo_apenas_entradas(self, db, usuario):
        from finances.services import create_transaction, calculate_balance
        create_transaction(usuario, {'valor': '1000.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'S1'})
        create_transaction(usuario, {'valor': '500.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'S2'})
        saldo = calculate_balance(usuario)
        assert saldo['total_entradas'] == Decimal('1500.00')
        assert saldo['saldo_liquido'] == Decimal('1500.00')


@pytest.mark.django_db
class TestDashboardRealtimeSignals:

    def test_create_transaction_dispara_push_dashboard(self, usuario, monkeypatch):
        chamadas = []

        class FakeLayer:
            def group_send(self, group_name, payload):
                chamadas.append((group_name, payload))

        monkeypatch.setattr('finances.signals._resolve_channel_layer', lambda: FakeLayer())
        monkeypatch.setattr('finances.signals.async_to_sync', lambda fn: fn)

        Transaction.objects.create(
            usuario=usuario,
            valor=Decimal('123.45'),
            data=date.today(),
            tipo='entrada',
            descricao='Transação WS',
        )

        assert len(chamadas) == 1
        group_name, payload = chamadas[0]
        assert group_name == f'dashboard_user_{usuario.id}'
        assert payload['type'] == 'dashboard.update'

    def test_update_transaction_nao_dispara_push_dashboard(self, usuario, monkeypatch):
        chamadas = []

        class FakeLayer:
            def group_send(self, group_name, payload):
                chamadas.append((group_name, payload))

        monkeypatch.setattr('finances.signals._resolve_channel_layer', lambda: FakeLayer())
        monkeypatch.setattr('finances.signals.async_to_sync', lambda fn: fn)

        tx = Transaction.objects.create(
            usuario=usuario,
            valor=Decimal('50.00'),
            data=date.today(),
            tipo='saida',
            descricao='Original',
        )
        chamadas.clear()

        tx.descricao = 'Atualizada'
        tx.save()

        assert chamadas == []

    def test_saldo_misto(self, db, usuario):
        from finances.services import create_transaction, calculate_balance
        create_transaction(usuario, {'valor': '2000.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'E'})
        create_transaction(usuario, {'valor': '600.00', 'data': date.today(), 'tipo': 'saida', 'descricao': 'S'})
        saldo = calculate_balance(usuario)
        assert saldo['total_entradas'] == Decimal('2000.00')
        assert saldo['total_saidas'] == Decimal('600.00')
        assert saldo['saldo_liquido'] == Decimal('1400.00')

    def test_saldo_isolado_por_usuario(self, db, usuario):
        from finances.services import create_transaction, calculate_balance
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro4', password='pass123')
        create_transaction(outro, {'valor': '9999.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'Outro'})
        saldo = calculate_balance(usuario)
        assert saldo['total_entradas'] == Decimal('0.00')


class TestSelectors:
    """Testes unitários para get_all_transactions e get_transaction_by_id."""

    def test_get_all_transactions_vazio(self, db, usuario):
        from finances.selectors import get_all_transactions
        qs = get_all_transactions(usuario)
        assert qs.count() == 0

    def test_get_all_transactions_retorna_somente_do_usuario(self, db, usuario):
        from finances.services import create_transaction
        from finances.selectors import get_all_transactions
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro5', password='pass123')
        create_transaction(usuario, {'valor': '100.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'Minha'})
        create_transaction(outro, {'valor': '200.00', 'data': date.today(), 'tipo': 'entrada', 'descricao': 'Dele'})
        qs = get_all_transactions(usuario)
        assert qs.count() == 1
        assert qs.first().descricao == 'Minha'

    def test_get_transaction_by_id_encontra(self, db, usuario):
        from finances.services import create_transaction
        from finances.selectors import get_transaction_by_id
        t = create_transaction(usuario, {'valor': '50.00', 'data': date.today(), 'tipo': 'saida', 'descricao': 'Busca'})
        encontrada = get_transaction_by_id(t.pk, usuario)
        assert encontrada.pk == t.pk

    def test_get_transaction_by_id_nao_encontrado_levanta_404(self, db, usuario):
        from django.http import Http404
        from finances.selectors import get_transaction_by_id
        with pytest.raises(Http404):
            get_transaction_by_id(99999, usuario)

    def test_get_transaction_by_id_outro_usuario_levanta_404(self, db, usuario):
        from django.http import Http404
        from finances.services import create_transaction
        from finances.selectors import get_transaction_by_id
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro6', password='pass123')
        t = create_transaction(outro, {'valor': '50.00', 'data': date.today(), 'tipo': 'saida', 'descricao': 'Dele'})
        with pytest.raises(Http404):
            get_transaction_by_id(t.pk, usuario)


# ============================================================
# Fase 4 — Integração HTTP (Views)
# ============================================================


@pytest.mark.django_db
class TestTransactionViews:

    def test_transaction_list_autenticado(self, client_autenticado):
        url = reverse('finances:transaction_list')
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_transaction_list_template_conecta_websocket(self, client_autenticado):
        url = reverse('finances:transaction_list')
        response = client_autenticado.get(url)
        assert response.status_code == 200
        assert b'ws-connect="/ws/dashboard/"' in response.content

    def test_transaction_list_requer_login(self, client):
        url = reverse('finances:transaction_list')
        response = client.get(url)
        assert response.status_code == 302
        assert '/accounts/login/' in response['Location']

    def test_transaction_create_get(self, client_autenticado):
        url = reverse('finances:transaction_create')
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_transaction_create_post_valido(self, client_autenticado, usuario):
        url = reverse('finances:transaction_create')
        data = {
            'valor': '100.00',
            'data': '2024-01-15',
            'tipo': 'entrada',
            'descricao': 'Salário Teste',
        }
        response = client_autenticado.post(url, data)
        assert response.status_code == 302
        assert Transaction.objects.filter(usuario=usuario, descricao='Salário Teste').exists()

    def test_transaction_create_post_invalido(self, client_autenticado):
        url = reverse('finances:transaction_create')
        response = client_autenticado.post(url, {'valor': '-50', 'descricao': ''})
        assert response.status_code == 200

    def test_transaction_update_get(self, client_autenticado, usuario):
        t = Transaction.objects.create(
            usuario=usuario, valor='50.00', data='2024-01-01',
            tipo='saida', descricao='Aluguel'
        )
        url = reverse('finances:transaction_update', kwargs={'pk': t.pk})
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_transaction_update_outro_usuario(self, client_autenticado):
        outro = User.objects.create_user(username='outro', password='pass123')
        t = Transaction.objects.create(
            usuario=outro, valor='30.00', data='2024-01-01',
            tipo='saida', descricao='Outra'
        )
        url = reverse('finances:transaction_update', kwargs={'pk': t.pk})
        response = client_autenticado.get(url)
        assert response.status_code == 404

    def test_transaction_delete_get(self, client_autenticado, usuario):
        t = Transaction.objects.create(
            usuario=usuario, valor='20.00', data='2024-01-01',
            tipo='saida', descricao='Excluir'
        )
        url = reverse('finances:transaction_delete', kwargs={'pk': t.pk})
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_transaction_delete_post(self, client_autenticado, usuario):
        t = Transaction.objects.create(
            usuario=usuario, valor='20.00', data='2024-01-01',
            tipo='saida', descricao='Excluir'
        )
        url = reverse('finances:transaction_delete', kwargs={'pk': t.pk})
        response = client_autenticado.post(url)
        assert response.status_code == 302
        assert not Transaction.objects.filter(pk=t.pk).exists()

    def test_transaction_delete_outro_usuario(self, client_autenticado):
        outro = User.objects.create_user(username='outro2', password='pass123')
        t = Transaction.objects.create(
            usuario=outro, valor='30.00', data='2024-01-01',
            tipo='saida', descricao='Protegida'
        )
        url = reverse('finances:transaction_delete', kwargs={'pk': t.pk})
        response = client_autenticado.post(url)
        assert response.status_code == 404
        assert Transaction.objects.filter(pk=t.pk).exists()


@pytest.mark.django_db
class TestCategoryViews:

    def test_category_list_autenticado(self, client_autenticado):
        url = reverse('finances:category_list')
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_category_list_requer_login(self, client):
        url = reverse('finances:category_list')
        response = client.get(url)
        assert response.status_code == 302
        assert '/accounts/login/' in response['Location']

    def test_category_create_get(self, client_autenticado):
        url = reverse('finances:category_create')
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_category_create_post_valido(self, client_autenticado, usuario):
        url = reverse('finances:category_create')
        data = {'nome': 'Alimentação', 'tipo': 'saida', 'keywords': ''}
        response = client_autenticado.post(url, data)
        assert response.status_code == 302
        assert Category.objects.filter(usuario=usuario, nome='Alimentação').exists()

    def test_category_update_outro_usuario(self, client_autenticado):
        outro = User.objects.create_user(username='terceiro', password='pass123')
        cat = Category.objects.create(
            usuario=outro, nome='CatOutro', tipo='saida'
        )
        url = reverse('finances:category_update', kwargs={'pk': cat.pk})
        response = client_autenticado.get(url)
        assert response.status_code == 404

    def test_category_delete_post(self, client_autenticado, usuario):
        cat = Category.objects.create(
            usuario=usuario, nome='ParaExcluir', tipo='saida'
        )
        url = reverse('finances:category_delete', kwargs={'pk': cat.pk})
        response = client_autenticado.post(url)
        assert response.status_code == 302
        assert not Category.objects.filter(pk=cat.pk).exists()


# ============================================================
# Fase 5 — Filtros na TransactionListView
# ============================================================


@pytest.mark.django_db
class TestTransactionListViewFiltros:

    def test_filtro_tipo_entrada(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Transaction
        import datetime
        Transaction.objects.create(usuario=usuario, valor='10.00', data=datetime.date(2024,1,1), tipo='entrada', descricao='Entrada')
        Transaction.objects.create(usuario=usuario, valor='20.00', data=datetime.date(2024,1,1), tipo='saida', descricao='Saída')
        url = reverse('finances:transaction_list') + '?tipo=entrada'
        response = client_autenticado.get(url)
        assert response.status_code == 200
        transacoes = list(response.context['transacoes'])
        assert all(t.tipo == 'entrada' for t in transacoes)
        assert len(transacoes) == 1

    def test_filtro_tipo_saida(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Transaction
        import datetime
        Transaction.objects.create(usuario=usuario, valor='10.00', data=datetime.date(2024,1,1), tipo='entrada', descricao='Entrada')
        Transaction.objects.create(usuario=usuario, valor='20.00', data=datetime.date(2024,1,1), tipo='saida', descricao='Saída')
        url = reverse('finances:transaction_list') + '?tipo=saida'
        response = client_autenticado.get(url)
        transacoes = list(response.context['transacoes'])
        assert all(t.tipo == 'saida' for t in transacoes)

    def test_filtro_descricao(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Transaction
        import datetime
        Transaction.objects.create(usuario=usuario, valor='50.00', data=datetime.date(2024,1,1), tipo='entrada', descricao='Salário Janeiro')
        Transaction.objects.create(usuario=usuario, valor='30.00', data=datetime.date(2024,1,1), tipo='saida', descricao='Supermercado')
        url = reverse('finances:transaction_list') + '?q=salário'
        response = client_autenticado.get(url)
        transacoes = list(response.context['transacoes'])
        assert len(transacoes) == 1
        assert transacoes[0].descricao == 'Salário Janeiro'

    def test_filtro_data_inicio(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Transaction
        import datetime
        Transaction.objects.create(usuario=usuario, valor='10.00', data=datetime.date(2024,1,1), tipo='entrada', descricao='Jan')
        Transaction.objects.create(usuario=usuario, valor='10.00', data=datetime.date(2024,3,1), tipo='entrada', descricao='Mar')
        url = reverse('finances:transaction_list') + '?data_inicio=2024-02-01'
        response = client_autenticado.get(url)
        transacoes = list(response.context['transacoes'])
        assert all(t.data >= datetime.date(2024, 2, 1) for t in transacoes)

    def test_filtros_no_contexto(self, client_autenticado):
        from django.urls import reverse
        url = reverse('finances:transaction_list') + '?tipo=entrada&q=teste'
        response = client_autenticado.get(url)
        assert response.context['filtros']['tipo'] == 'entrada'
        assert response.context['filtros']['q'] == 'teste'


# ============================================================
# Fase 6 — Goal Model, Services e Views
# ============================================================


@pytest.mark.django_db
class TestGoalModel:
    def test_percentual_concluido_zero_sem_progresso(self, usuario):
        from finances.models import Goal
        meta = Goal.objects.create(usuario=usuario, titulo='T', valor_alvo='1000.00')
        assert meta.percentual_concluido == 0

    def test_percentual_concluido_50_percent(self, usuario):
        from finances.models import Goal
        meta = Goal.objects.create(usuario=usuario, titulo='T', valor_alvo='1000.00', valor_atual='500.00')
        assert meta.percentual_concluido == 50

    def test_percentual_concluido_nao_excede_100(self, usuario):
        from finances.models import Goal
        meta = Goal.objects.create(usuario=usuario, titulo='T', valor_alvo='100.00', valor_atual='200.00')
        assert meta.percentual_concluido == 100

    def test_saldo_restante(self, usuario):
        from finances.models import Goal
        from decimal import Decimal
        meta = Goal.objects.create(usuario=usuario, titulo='T', valor_alvo='1000.00', valor_atual='300.00')
        assert meta.saldo_restante == Decimal('700.00')

    def test_saldo_restante_zero_quando_concluido(self, usuario):
        from finances.models import Goal
        from decimal import Decimal
        meta = Goal.objects.create(usuario=usuario, titulo='T', valor_alvo='500.00', valor_atual='600.00')
        assert meta.saldo_restante == Decimal('0.00')


@pytest.mark.django_db
class TestGoalServices:
    def test_create_goal_basico(self, usuario):
        from finances.services import create_goal
        from finances.models import Goal
        meta = create_goal(usuario, {'titulo': 'Viagem', 'valor_alvo': '5000.00'})
        assert Goal.objects.filter(pk=meta.pk, usuario=usuario).exists()

    def test_create_goal_valor_zero_raises(self, usuario):
        from finances.services import create_goal
        from django.core.exceptions import ValidationError
        with pytest.raises(ValidationError):
            create_goal(usuario, {'titulo': 'X', 'valor_alvo': '0'})

    def test_update_goal(self, usuario):
        from finances.services import create_goal, update_goal
        meta = create_goal(usuario, {'titulo': 'Viagem', 'valor_alvo': '1000.00'})
        update_goal(meta.pk, usuario, {'titulo': 'Viagem Internacional'})
        meta.refresh_from_db()
        assert meta.titulo == 'Viagem Internacional'

    def test_delete_goal(self, usuario):
        from finances.services import create_goal, delete_goal
        from finances.models import Goal
        meta = create_goal(usuario, {'titulo': 'Del', 'valor_alvo': '100.00'})
        delete_goal(meta.pk, usuario)
        assert not Goal.objects.filter(pk=meta.pk).exists()

    def test_add_progress_incrementa(self, usuario):
        from finances.services import create_goal, add_progress_to_goal
        from decimal import Decimal
        meta = create_goal(usuario, {'titulo': 'G', 'valor_alvo': '1000.00'})
        add_progress_to_goal(meta.pk, usuario, '300.00')
        meta.refresh_from_db()
        assert meta.valor_atual == Decimal('300.00')

    def test_add_progress_nao_excede_alvo(self, usuario):
        from finances.services import create_goal, add_progress_to_goal
        from decimal import Decimal
        meta = create_goal(usuario, {'titulo': 'G', 'valor_alvo': '500.00'})
        add_progress_to_goal(meta.pk, usuario, '700.00')
        meta.refresh_from_db()
        assert meta.valor_atual == Decimal('500.00')

    def test_add_progress_valor_negativo_raises(self, usuario):
        from finances.services import create_goal, add_progress_to_goal
        from django.core.exceptions import ValidationError
        meta = create_goal(usuario, {'titulo': 'G', 'valor_alvo': '100.00'})
        with pytest.raises(ValidationError):
            add_progress_to_goal(meta.pk, usuario, '-10')


@pytest.mark.django_db
class TestAccountViews:
    def test_account_list_200(self, client_autenticado):
        from django.urls import reverse
        response = client_autenticado.get(reverse('finances:account_list'))
        assert response.status_code == 200

    def test_account_list_requer_login(self, client):
        from django.urls import reverse
        response = client.get(reverse('finances:account_list'))
        assert response.status_code == 302

    def test_account_create_get(self, client_autenticado):
        from django.urls import reverse
        response = client_autenticado.get(reverse('finances:account_create'))
        assert response.status_code == 200

    def test_account_create_post_valido(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Account
        response = client_autenticado.post(reverse('finances:account_create'), {
            'nome': 'Nubank', 'tipo': 'conta_corrente'
        })
        assert response.status_code == 302
        assert Account.objects.filter(usuario=usuario, nome='Nubank').exists()

    def test_account_delete_outro_usuario(self, client_autenticado):
        from django.urls import reverse
        from finances.models import Account
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro_acc', password='pass')
        conta = Account.objects.create(usuario=outro, nome='Conta Alheia', tipo='carteira')
        response = client_autenticado.post(reverse('finances:account_delete', kwargs={'pk': conta.pk}))
        assert response.status_code == 404


@pytest.mark.django_db
class TestGoalViews:
    def test_goal_list_200(self, client_autenticado):
        from django.urls import reverse
        response = client_autenticado.get(reverse('finances:goal_list'))
        assert response.status_code == 200

    def test_goal_list_requer_login(self, client):
        from django.urls import reverse
        response = client.get(reverse('finances:goal_list'))
        assert response.status_code == 302

    def test_goal_create_get(self, client_autenticado):
        from django.urls import reverse
        response = client_autenticado.get(reverse('finances:goal_create'))
        assert response.status_code == 200

    def test_goal_create_post_valido(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Goal
        response = client_autenticado.post(reverse('finances:goal_create'), {
            'titulo': 'Reserva', 'valor_alvo': '10000.00'
        })
        assert response.status_code == 302
        assert Goal.objects.filter(usuario=usuario, titulo='Reserva').exists()

    def test_goal_add_progress_post(self, client_autenticado, usuario):
        from django.urls import reverse
        from finances.models import Goal
        from decimal import Decimal
        meta = Goal.objects.create(usuario=usuario, titulo='G', valor_alvo='1000.00')
        response = client_autenticado.post(
            reverse('finances:goal_add_progress', kwargs={'pk': meta.pk}),
            {'valor': '200.00'}
        )
        assert response.status_code == 302
        meta.refresh_from_db()
        assert meta.valor_atual == Decimal('200.00')

    def test_goal_delete_outro_usuario(self, client_autenticado):
        from django.urls import reverse
        from finances.models import Goal
        from django.contrib.auth.models import User
        outro = User.objects.create_user(username='outro_goal', password='pass')
        meta = Goal.objects.create(usuario=outro, titulo='Alheia', valor_alvo='100.00')
        response = client_autenticado.post(reverse('finances:goal_delete', kwargs={'pk': meta.pk}))
        assert response.status_code == 404
        assert Goal.objects.filter(pk=meta.pk).exists()


# ---------------------------------------------------------------------------
# Fase 7 — Testes de get_account_balance e CBVs separadas
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestGetAccountBalance:
    """Testes unitários para finances.selectors.get_account_balance."""

    def _make_account(self, usuario, tipo='conta_corrente'):
        return Account.objects.create(nome=f'Conta {tipo}', tipo=tipo, usuario=usuario)

    def _make_transaction(self, usuario, conta, tipo, valor):
        cat, _ = Category.objects.get_or_create(
            nome='Avulsa', usuario=usuario, defaults={'tipo': tipo}
        )
        return Transaction.objects.create(
            usuario=usuario, conta=conta, categoria=cat,
            tipo=tipo, valor=valor, data=date.today(), descricao='test',
        )

    def test_conta_sem_transacoes_retorna_zero(self, db, usuario):
        from finances.selectors import get_account_balance
        conta = self._make_account(usuario)
        assert get_account_balance(conta.pk, usuario) == Decimal('0')

    def test_conta_corrente_entradas_menos_saidas(self, db, usuario):
        from finances.selectors import get_account_balance
        conta = self._make_account(usuario, tipo='conta_corrente')
        self._make_transaction(usuario, conta, 'entrada', Decimal('1000.00'))
        self._make_transaction(usuario, conta, 'saida', Decimal('300.00'))
        assert get_account_balance(conta.pk, usuario) == Decimal('700.00')

    def test_carteira_saldo_positivo(self, db, usuario):
        from finances.selectors import get_account_balance
        conta = self._make_account(usuario, tipo='carteira')
        self._make_transaction(usuario, conta, 'entrada', Decimal('500.00'))
        assert get_account_balance(conta.pk, usuario) == Decimal('500.00')

    def test_cartao_credito_retorna_soma_saidas(self, db, usuario):
        from finances.selectors import get_account_balance
        cartao = self._make_account(usuario, tipo='cartao_credito')
        self._make_transaction(usuario, cartao, 'saida', Decimal('200.00'))
        self._make_transaction(usuario, cartao, 'saida', Decimal('150.00'))
        assert get_account_balance(cartao.pk, usuario) == Decimal('350.00')

    def test_cartao_sem_transacoes_fatura_zero(self, db, usuario):
        from finances.selectors import get_account_balance
        cartao = self._make_account(usuario, tipo='cartao_credito')
        assert get_account_balance(cartao.pk, usuario) == Decimal('0')

    def test_nao_vaza_entre_usuarios(self, db, usuario):
        from finances.selectors import get_account_balance
        from django.http import Http404
        outro = User.objects.create_user(username='outro_bal', password='pass')
        conta_outro = self._make_account(outro, tipo='conta_corrente')
        with pytest.raises(Http404):
            get_account_balance(conta_outro.pk, usuario)

    def test_saldo_isolado_por_conta(self, db, usuario):
        from finances.selectors import get_account_balance
        conta_a = self._make_account(usuario, tipo='conta_corrente')
        conta_b = Account.objects.create(nome='Conta B', tipo='conta_corrente', usuario=usuario)
        cat = Category.objects.get_or_create(nome='Avulsa', tipo='entrada', usuario=usuario)[0]
        Transaction.objects.create(
            usuario=usuario, conta=conta_a, categoria=cat,
            tipo='entrada', valor=Decimal('1000.00'), data=date.today(), descricao='A',
        )
        # conta_b sem transações deve retornar 0
        assert get_account_balance(conta_b.pk, usuario) == Decimal('0')
        assert get_account_balance(conta_a.pk, usuario) == Decimal('1000.00')


@pytest.mark.django_db
class TestContaCorrenteViews:
    """Testes HTTP para ContaCorrenteListView e ContaCorrenteCreateView."""

    def test_lista_requer_login(self, client):
        response = client.get(reverse('finances:conta_corrente_list'))
        assert response.status_code == 302

    def test_lista_retorna_200_autenticado(self, client_autenticado):
        response = client_autenticado.get(reverse('finances:conta_corrente_list'))
        assert response.status_code == 200

    def test_lista_exibe_somente_conta_corrente_e_carteira(self, client_autenticado, usuario):
        Account.objects.create(nome='Bradesco', tipo='conta_corrente', usuario=usuario)
        Account.objects.create(nome='Carteira', tipo='carteira', usuario=usuario)
        Account.objects.create(nome='Nubank', tipo='cartao_credito', usuario=usuario)
        response = client_autenticado.get(reverse('finances:conta_corrente_list'))
        assert response.status_code == 200
        contas = response.context['contas']
        tipos = {c.tipo for c in contas}
        assert 'cartao_credito' not in tipos
        assert len(contas) == 2

    def test_create_get_retorna_200(self, client_autenticado):
        response = client_autenticado.get(reverse('finances:conta_corrente_create'))
        assert response.status_code == 200

    def test_create_post_cria_conta_corrente(self, client_autenticado, usuario):
        response = client_autenticado.post(
            reverse('finances:conta_corrente_create'),
            {'nome': 'Itaú', 'tipo': 'conta_corrente'},
        )
        assert response.status_code == 302
        assert Account.objects.filter(usuario=usuario, nome='Itaú', tipo='conta_corrente').exists()

    def test_delete_outro_usuario_retorna_404(self, client_autenticado):
        outro = User.objects.create_user(username='outro_cc', password='pass')
        conta = Account.objects.create(nome='Alheia', tipo='conta_corrente', usuario=outro)
        response = client_autenticado.post(
            reverse('finances:conta_corrente_delete', kwargs={'pk': conta.pk})
        )
        assert response.status_code == 404

    def test_context_tem_contas_com_saldo(self, client_autenticado, usuario):
        conta = Account.objects.create(nome='BB', tipo='conta_corrente', usuario=usuario)
        response = client_autenticado.get(reverse('finances:conta_corrente_list'))
        assert 'contas_com_saldo' in response.context
        pks = [c.pk for c, _, _ in response.context['contas_com_saldo']]
        assert conta.pk in pks


@pytest.mark.django_db
class TestCartaoCreditoViews:
    """Testes HTTP para CartaoCreditoListView e CartaoCreditoCreateView."""

    def test_lista_requer_login(self, client):
        response = client.get(reverse('finances:cartao_credito_list'))
        assert response.status_code == 302

    def test_lista_retorna_200_autenticado(self, client_autenticado):
        response = client_autenticado.get(reverse('finances:cartao_credito_list'))
        assert response.status_code == 200

    def test_lista_exibe_somente_cartao_credito(self, client_autenticado, usuario):
        Account.objects.create(nome='Bradesco', tipo='conta_corrente', usuario=usuario)
        Account.objects.create(nome='Nubank', tipo='cartao_credito', usuario=usuario)
        response = client_autenticado.get(reverse('finances:cartao_credito_list'))
        cartoes = response.context['cartoes']
        tipos = {c.tipo for c in cartoes}
        assert tipos == {'cartao_credito'}

    def test_create_get_retorna_200(self, client_autenticado):
        response = client_autenticado.get(reverse('finances:cartao_credito_create'))
        assert response.status_code == 200

    def test_create_post_cria_cartao(self, client_autenticado, usuario):
        response = client_autenticado.post(
            reverse('finances:cartao_credito_create'),
            {'nome': 'Inter', 'tipo': 'cartao_credito'},
        )
        assert response.status_code == 302
        assert Account.objects.filter(usuario=usuario, nome='Inter', tipo='cartao_credito').exists()

    def test_delete_outro_usuario_retorna_404(self, client_autenticado):
        outro = User.objects.create_user(username='outro_cartao', password='pass')
        cartao = Account.objects.create(nome='Alheio', tipo='cartao_credito', usuario=outro)
        response = client_autenticado.post(
            reverse('finances:cartao_credito_delete', kwargs={'pk': cartao.pk})
        )
        assert response.status_code == 404

    def test_context_tem_cartoes_com_fatura(self, client_autenticado, usuario):
        cartao = Account.objects.create(nome='XP', tipo='cartao_credito', usuario=usuario)
        response = client_autenticado.get(reverse('finances:cartao_credito_list'))
        assert 'cartoes_com_fatura' in response.context
        pks = [c.pk for c, _ in response.context['cartoes_com_fatura']]
        assert cartao.pk in pks


# ============================================================
# Relatórios — Selector, Services, Form, View
# ============================================================


@pytest.fixture
def transacoes_relatorio(usuario):
    """Cria transações de teste para relatórios."""
    cat_alim = Category.objects.create(
        nome='Alimentação', tipo='saida', usuario=usuario,
    )
    cat_transp = Category.objects.create(
        nome='Transporte', tipo='saida', usuario=usuario,
    )
    conta = Account.objects.create(
        nome='Nubank', tipo='conta_corrente', usuario=usuario,
    )
    Transaction.objects.create(
        usuario=usuario, valor='3000.00', data=date(2026, 1, 10),
        tipo='entrada', descricao='Salário', conta=conta,
    )
    Transaction.objects.create(
        usuario=usuario, valor='500.00', data=date(2026, 1, 15),
        tipo='saida', descricao='Supermercado', categoria=cat_alim, conta=conta,
    )
    Transaction.objects.create(
        usuario=usuario, valor='200.00', data=date(2026, 1, 20),
        tipo='saida', descricao='Uber', categoria=cat_transp, conta=conta,
    )
    # Fora do período (fevereiro)
    Transaction.objects.create(
        usuario=usuario, valor='100.00', data=date(2026, 2, 5),
        tipo='saida', descricao='Fora do período', categoria=cat_alim, conta=conta,
    )
    return {
        'cat_alim': cat_alim,
        'cat_transp': cat_transp,
        'conta': conta,
    }


@pytest.mark.django_db
class TestGetReportData:

    def test_resumo_totais(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert data['resumo']['total_entradas'] == Decimal('3000.00')
        assert data['resumo']['total_saidas'] == Decimal('700.00')
        assert data['resumo']['saldo_liquido'] == Decimal('2300.00')
        assert data['resumo']['num_transacoes'] == 3

    def test_resumo_maior_entrada_saida(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert data['resumo']['maior_entrada'] == Decimal('3000.00')
        assert data['resumo']['maior_saida'] == Decimal('500.00')

    def test_periodo(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert data['periodo']['inicio'] == date(2026, 1, 1)
        assert data['periodo']['fim'] == date(2026, 1, 31)

    def test_transacoes_filtradas_por_periodo(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert data['transacoes'].count() == 3

    def test_por_categoria(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        categorias = {c['categoria']: c for c in data['por_categoria']}
        assert 'Alimentação' in categorias
        assert 'Transporte' in categorias
        assert categorias['Alimentação']['total'] == Decimal('500.00')
        assert categorias['Transporte']['total'] == Decimal('200.00')

    def test_percentuais_somam_100(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2026, 1, 1), date(2026, 1, 31))
        total_pct = sum(c['percentual'] for c in data['por_categoria'])
        assert abs(total_pct - 100.0) < 0.1

    def test_ownership_isolado(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        outro = User.objects.create_user(username='outro_rel', password='pass123')
        data = get_report_data(outro, date(2026, 1, 1), date(2026, 1, 31))
        assert data['resumo']['num_transacoes'] == 0
        assert data['resumo']['total_entradas'] == Decimal('0.00')

    def test_sem_transacoes_no_periodo(self, usuario, transacoes_relatorio):
        from finances.selectors import get_report_data
        data = get_report_data(usuario, date(2025, 1, 1), date(2025, 1, 31))
        assert data['resumo']['num_transacoes'] == 0
        assert data['resumo']['maior_entrada'] is None
        assert data['resumo']['maior_saida'] is None


@pytest.mark.django_db
class TestFormatBrl:

    def test_format_brl_positivo(self):
        from finances.services import _format_brl
        assert _format_brl(Decimal('1234.56')) == 'R$ 1.234,56'

    def test_format_brl_zero(self):
        from finances.services import _format_brl
        assert _format_brl(Decimal('0.00')) == 'R$ 0,00'

    def test_format_brl_negativo(self):
        from finances.services import _format_brl
        assert _format_brl(Decimal('-500.00')) == '-R$ 500,00'

    def test_format_brl_none(self):
        from finances.services import _format_brl
        assert _format_brl(None) == 'R$ 0,00'

    def test_format_brl_grande(self):
        from finances.services import _format_brl
        assert _format_brl(Decimal('1000000.00')) == 'R$ 1.000.000,00'


@pytest.mark.django_db
class TestGeneratePdfReport:

    def test_pdf_retorna_bytes(self, usuario, transacoes_relatorio):
        from finances.services import generate_pdf_report
        content = generate_pdf_report(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_pdf_header_valido(self, usuario, transacoes_relatorio):
        from finances.services import generate_pdf_report
        content = generate_pdf_report(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert content[:5] == b'%PDF-'

    def test_pdf_sem_transacoes(self, usuario):
        from finances.services import generate_pdf_report
        content = generate_pdf_report(usuario, date(2025, 1, 1), date(2025, 1, 31))
        assert content[:5] == b'%PDF-'


@pytest.mark.django_db
class TestGenerateExcelReport:

    def test_excel_retorna_bytes(self, usuario, transacoes_relatorio):
        from finances.services import generate_excel_report
        content = generate_excel_report(usuario, date(2026, 1, 1), date(2026, 1, 31))
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_excel_abas(self, usuario, transacoes_relatorio):
        from finances.services import generate_excel_report
        from openpyxl import load_workbook
        content = generate_excel_report(usuario, date(2026, 1, 1), date(2026, 1, 31))
        wb = load_workbook(io.BytesIO(content))
        assert 'Resumo' in wb.sheetnames
        assert 'Transações' in wb.sheetnames
        assert 'Por Categoria' in wb.sheetnames

    def test_excel_aba_transacoes_com_dados(self, usuario, transacoes_relatorio):
        from finances.services import generate_excel_report
        from openpyxl import load_workbook
        content = generate_excel_report(usuario, date(2026, 1, 1), date(2026, 1, 31))
        wb = load_workbook(io.BytesIO(content))
        ws = wb['Transações']
        # Header + 3 transações
        assert ws.max_row == 4

    def test_excel_sem_transacoes(self, usuario):
        from finances.services import generate_excel_report
        from openpyxl import load_workbook
        content = generate_excel_report(usuario, date(2025, 1, 1), date(2025, 1, 31))
        wb = load_workbook(io.BytesIO(content))
        ws = wb['Transações']
        assert ws.max_row == 1  # Apenas header


@pytest.mark.django_db
class TestReportFilterForm:

    def test_form_valido(self):
        from finances.forms import ReportFilterForm
        form = ReportFilterForm(data={
            'data_inicio': '2026-01-01',
            'data_fim': '2026-01-31',
            'formato': 'pdf',
        })
        assert form.is_valid()

    def test_form_datas_invertidas(self):
        from finances.forms import ReportFilterForm
        form = ReportFilterForm(data={
            'data_inicio': '2026-02-01',
            'data_fim': '2026-01-01',
            'formato': 'pdf',
        })
        assert not form.is_valid()

    def test_form_formato_excel(self):
        from finances.forms import ReportFilterForm
        form = ReportFilterForm(data={
            'data_inicio': '2026-01-01',
            'data_fim': '2026-01-31',
            'formato': 'excel',
        })
        assert form.is_valid()

    def test_form_formato_invalido(self):
        from finances.forms import ReportFilterForm
        form = ReportFilterForm(data={
            'data_inicio': '2026-01-01',
            'data_fim': '2026-01-31',
            'formato': 'csv',
        })
        assert not form.is_valid()


@pytest.mark.django_db
class TestReportView:

    def test_report_get_autenticado(self, client_autenticado):
        url = reverse('finances:report')
        response = client_autenticado.get(url)
        assert response.status_code == 200

    def test_report_requer_login(self, client):
        url = reverse('finances:report')
        response = client.get(url)
        assert response.status_code == 302
        assert '/accounts/login/' in response['Location']

    def test_report_pdf_download(self, client_autenticado, usuario, transacoes_relatorio):
        url = reverse('finances:report')
        response = client_autenticado.post(url, {
            'data_inicio': '2026-01-01',
            'data_fim': '2026-01-31',
            'formato': 'pdf',
        })
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/pdf'
        assert 'attachment' in response['Content-Disposition']
        assert response.content[:5] == b'%PDF-'

    def test_report_excel_download(self, client_autenticado, usuario, transacoes_relatorio):
        url = reverse('finances:report')
        response = client_autenticado.post(url, {
            'data_inicio': '2026-01-01',
            'data_fim': '2026-01-31',
            'formato': 'excel',
        })
        assert response.status_code == 200
        assert 'spreadsheetml' in response['Content-Type']
        assert 'attachment' in response['Content-Disposition']

    def test_report_datas_invertidas_rejeita(self, client_autenticado):
        url = reverse('finances:report')
        response = client_autenticado.post(url, {
            'data_inicio': '2026-02-01',
            'data_fim': '2026-01-01',
            'formato': 'pdf',
        })
        assert response.status_code == 200  # Re-renders form with errors
