import csv
import datetime
import io
import logging
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction as db_transaction
from django.db.models import Sum

from .models import BudgetAlertEvent, Category, MonthlyBudgetConfig, Transaction

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CRUD de Transações
# ---------------------------------------------------------------------------

def create_transaction(user, data: dict) -> 'Transaction':
    """
    Cria uma nova transação validando os dados.

    Args:
        user: instância de User
        data: dict com chaves valor, data, tipo, descricao, categoria_id (opt), conta_id (opt)

    Returns:
        Transaction criada

    Raises:
        ValidationError: se os dados forem inválidos
    """
    valor = data.get('valor')
    if valor is None or Decimal(str(valor)) <= 0:
        raise ValidationError('O valor da transação deve ser maior que zero.')

    tipo = data.get('tipo')
    if tipo not in ('entrada', 'saida'):
        raise ValidationError('O tipo deve ser "entrada" ou "saida".')

    from .models import Account

    categoria = None
    if data.get('categoria_id'):
        try:
            categoria = Category.objects.get(pk=data['categoria_id'], usuario=user)
        except Category.DoesNotExist:
            raise ValidationError('Categoria não encontrada ou não pertence ao usuário.')

    conta = None
    if data.get('conta_id'):
        try:
            conta = Account.objects.get(pk=data['conta_id'], usuario=user)
        except Account.DoesNotExist:
            raise ValidationError('Conta não encontrada ou não pertence ao usuário.')

    transacao = Transaction(
        valor=Decimal(str(valor)),
        data=data['data'],
        tipo=tipo,
        descricao=data.get('descricao', ''),
        categoria=categoria,
        conta=conta,
        usuario=user,
    )
    transacao.full_clean()
    transacao.save()
    return transacao


def update_transaction(transaction_id: int, user, data: dict) -> 'Transaction':
    """
    Atualiza uma transação existente com escopo por usuário.

    Raises:
        Transaction.DoesNotExist: se não encontrada ou não pertencer ao usuário
        ValidationError: se dados inválidos
    """
    from .models import Account

    transacao = Transaction.objects.get(pk=transaction_id, usuario=user)

    if 'valor' in data:
        valor = Decimal(str(data['valor']))
        if valor <= 0:
            raise ValidationError('O valor da transação deve ser maior que zero.')
        transacao.valor = valor

    if 'data' in data:
        transacao.data = data['data']

    if 'tipo' in data:
        if data['tipo'] not in ('entrada', 'saida'):
            raise ValidationError('O tipo deve ser "entrada" ou "saida".')
        transacao.tipo = data['tipo']

    if 'descricao' in data:
        transacao.descricao = data['descricao']

    if 'categoria_id' in data:
        if data['categoria_id']:
            try:
                transacao.categoria = Category.objects.get(
                    pk=data['categoria_id'], usuario=user
                )
            except Category.DoesNotExist:
                raise ValidationError('Categoria não encontrada ou não pertence ao usuário.')
        else:
            transacao.categoria = None

    if 'conta_id' in data:
        if data['conta_id']:
            try:
                transacao.conta = Account.objects.get(pk=data['conta_id'], usuario=user)
            except Account.DoesNotExist:
                raise ValidationError('Conta não encontrada ou não pertence ao usuário.')
        else:
            transacao.conta = None

    transacao.full_clean()
    transacao.save()
    return transacao


def delete_transaction(transaction_id: int, user) -> None:
    """
    Remove uma transação com escopo por usuário.

    Raises:
        Transaction.DoesNotExist: se não encontrada ou não pertencer ao usuário
    """
    transacao = Transaction.objects.get(pk=transaction_id, usuario=user)
    transacao.delete()


# ---------------------------------------------------------------------------
# Cálculo de saldo
# ---------------------------------------------------------------------------

def calculate_balance(user) -> dict:
    """
    Calcula o saldo financeiro do usuário.

    Returns:
        dict com total_entradas, total_saidas e saldo_liquido (todos Decimal)
    """
    total_entradas = (
        Transaction.objects.filter(usuario=user, tipo='entrada')
        .aggregate(total=Sum('valor'))['total']
        or Decimal('0.00')
    )

    total_saidas = (
        Transaction.objects.filter(usuario=user, tipo='saida')
        .aggregate(total=Sum('valor'))['total']
        or Decimal('0.00')
    )

    return {
        'total_entradas': total_entradas,
        'total_saidas': total_saidas,
        'saldo_liquido': total_entradas - total_saidas,
    }


# ---------------------------------------------------------------------------
# Auto-categorização
# ---------------------------------------------------------------------------

def _auto_categorize(descricao: str, user):
    """
    Cruza a descrição da transação com as keywords das categorias do usuário.
    Retorna a categoria correspondente ou a categoria 'Avulsa' como fallback.
    """
    descricao_lower = descricao.lower()
    categorias = Category.objects.filter(usuario=user).exclude(keywords='')

    for categoria in categorias:
        for keyword in categoria.get_keywords_list():
            if keyword and keyword in descricao_lower:
                return categoria

    # Fallback: categoria 'Avulsa'
    avulsa, _ = Category.objects.get_or_create(
        nome='Avulsa',
        usuario=user,
        defaults={'tipo': 'saida', 'keywords': ''},
    )
    return avulsa


# ---------------------------------------------------------------------------
# Parsers por formato
# ---------------------------------------------------------------------------

def _parse_csv(file_content: str) -> list[dict]:
    """Lê CSV e retorna lista de dicts com chaves: data, descricao, valor, tipo."""
    registros = []
    reader = csv.DictReader(io.StringIO(file_content))
    for row in reader:
        try:
            valor_str = row.get('valor', row.get('Valor', '0')).replace(',', '.')
            valor = abs(Decimal(valor_str))
            tipo = 'entrada' if Decimal(valor_str) >= 0 else 'saida'
            registros.append({
                'data': row.get('data', row.get('Data', '')),
                'descricao': row.get('descricao', row.get('Descricao', row.get('Histórico', ''))),
                'valor': valor,
                'tipo': tipo,
            })
        except (InvalidOperation, KeyError) as exc:
            logger.warning('Linha CSV ignorada: %s — %s', row, exc)
    return registros


def _parse_ofx(file_content: bytes) -> list[dict]:
    """Lê OFX e retorna lista de dicts padronizados."""
    try:
        from ofxparse import OfxParser

        ofx = OfxParser.parse(io.BytesIO(file_content))
        registros = []
        for account in [ofx.account] if hasattr(ofx, 'account') else []:
            for stmt in [account.statement] if hasattr(account, 'statement') else []:
                for t in stmt.transactions:
                    valor = abs(Decimal(str(t.amount)))
                    tipo = 'entrada' if t.amount >= 0 else 'saida'
                    registros.append({
                        'data': t.date.date() if hasattr(t.date, 'date') else t.date,
                        'descricao': t.memo or t.payee or '',
                        'valor': valor,
                        'tipo': tipo,
                    })
        return registros
    except Exception as exc:  # noqa: BLE001
        logger.error('Erro ao processar OFX: %s', exc)
        return []


def _parse_pdf(file_content: bytes) -> list[dict]:
    """Lê PDF bancário e retorna lista de dicts padronizados (heurística básica)."""
    try:
        import pdfplumber

        registros = []
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or len(row) < 3:
                            continue
                        try:
                            valor_str = str(row[-1]).replace(',', '.').replace('R$', '').strip()
                            valor = abs(Decimal(valor_str))
                            tipo = 'saida' if Decimal(valor_str) < 0 else 'entrada'
                            registros.append({
                                'data': str(row[0]).strip(),
                                'descricao': str(row[1]).strip() if len(row) > 1 else '',
                                'valor': valor,
                                'tipo': tipo,
                            })
                        except (InvalidOperation, IndexError) as exc:
                            logger.warning('Linha PDF ignorada: %s — %s', row, exc)
        return registros
    except Exception as exc:  # noqa: BLE001
        logger.error('Erro ao processar PDF: %s', exc)
        return []


# ---------------------------------------------------------------------------
# Serviço principal de importação
# ---------------------------------------------------------------------------

@db_transaction.atomic
def process_bank_statement_import(file, user, account) -> dict:
    """
    Processa extrato bancário (CSV, OFX ou PDF), auto-categoriza e salva as transações.

    Args:
        file: objeto de arquivo django (InMemoryUploadedFile ou similar)
        user: instância de User
        account: instância de Account

    Returns:
        dict com 'criadas' (int), 'ignoradas' (int) e 'erros' (list)
    """
    nome = file.name.lower()
    conteudo = file.read()

    if nome.endswith('.csv'):
        registros = _parse_csv(conteudo.decode('utf-8', errors='ignore'))
    elif nome.endswith('.ofx'):
        registros = _parse_ofx(conteudo)
    elif nome.endswith('.pdf'):
        registros = _parse_pdf(conteudo)
    else:
        return {'criadas': 0, 'ignoradas': 0, 'erros': [f'Formato não suportado: {nome}']}

    criadas = 0
    ignoradas = 0
    erros = []

    for reg in registros:
        try:
            import datetime

            data = reg['data']
            if isinstance(data, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        data = datetime.datetime.strptime(data, fmt).date()
                        break
                    except ValueError:
                        continue
                if isinstance(data, str):
                    ignoradas += 1
                    erros.append(f'Data inválida: {reg["data"]}')
                    continue

            categoria = _auto_categorize(reg['descricao'], user)

            Transaction.objects.create(
                valor=reg['valor'],
                data=data,
                tipo=reg['tipo'],
                descricao=reg['descricao'],
                categoria=categoria,
                conta=account,
                usuario=user,
            )
            criadas += 1
        except Exception as exc:  # noqa: BLE001
            ignoradas += 1
            erros.append(str(exc))
            logger.error('Erro ao salvar transação %s: %s', reg, exc)

    return {'criadas': criadas, 'ignoradas': ignoradas, 'erros': erros}


# ---------------------------------------------------------------------------
# CRUD de Metas
# ---------------------------------------------------------------------------

def create_goal(user, data: dict) -> 'Goal':
    """Cria uma nova meta financeira."""
    from .models import Goal
    valor_alvo = data.get('valor_alvo')
    if not valor_alvo or Decimal(str(valor_alvo)) <= 0:
        raise ValidationError('O valor alvo deve ser maior que zero.')

    categoria = None
    if data.get('categoria_id'):
        try:
            categoria = Category.objects.get(pk=data['categoria_id'], usuario=user)
        except Category.DoesNotExist:
            raise ValidationError('Categoria não encontrada ou não pertence ao usuário.')

    meta = Goal(
        titulo=data.get('titulo', ''),
        valor_alvo=Decimal(str(valor_alvo)),
        prazo=data.get('prazo'),
        categoria=categoria,
        usuario=user,
    )
    meta.full_clean()
    meta.save()
    return meta


def update_goal(goal_id: int, user, data: dict) -> 'Goal':
    """Atualiza uma meta existente com escopo por usuário."""
    from .models import Goal
    meta = Goal.objects.get(pk=goal_id, usuario=user)

    if 'titulo' in data:
        meta.titulo = data['titulo']
    if 'valor_alvo' in data and data['valor_alvo']:
        meta.valor_alvo = Decimal(str(data['valor_alvo']))
    if 'prazo' in data:
        meta.prazo = data.get('prazo')
    if 'categoria_id' in data:
        if data['categoria_id']:
            try:
                meta.categoria = Category.objects.get(pk=data['categoria_id'], usuario=user)
            except Category.DoesNotExist:
                raise ValidationError('Categoria não encontrada.')
        else:
            meta.categoria = None

    meta.full_clean()
    meta.save()
    return meta


def delete_goal(goal_id: int, user) -> None:
    """Remove uma meta com escopo por usuário."""
    from .models import Goal
    Goal.objects.filter(pk=goal_id, usuario=user).delete()


def add_progress_to_goal(goal_id: int, user, valor) -> 'Goal':
    """
    Incrementa o valor_atual da meta.
    Não permite exceder o valor_alvo.
    """
    from .models import Goal
    meta = Goal.objects.get(pk=goal_id, usuario=user)
    incremento = Decimal(str(valor))
    if incremento <= 0:
        raise ValidationError('O valor de progresso deve ser maior que zero.')
    meta.valor_atual = min(meta.valor_atual + incremento, meta.valor_alvo)
    meta.save(update_fields=['valor_atual'])
    return meta


# ---------------------------------------------------------------------------
# Orçamento Mensal
# ---------------------------------------------------------------------------

def upsert_monthly_budget_config(user, data: dict) -> 'MonthlyBudgetConfig':
    renda_mensal = Decimal(str(data['renda_mensal']))
    limite_percentual = Decimal(str(data['limite_percentual']))

    if renda_mensal <= 0:
        raise ValidationError('A renda mensal deve ser maior que zero.')
    if limite_percentual <= 0 or limite_percentual > 100:
        raise ValidationError('O percentual do limite deve estar entre 1 e 100.')

    config, _ = MonthlyBudgetConfig.objects.update_or_create(
        usuario=user,
        defaults={
            'renda_mensal': renda_mensal,
            'limite_percentual': limite_percentual,
            'alertas_ativos': bool(data.get('alertas_ativos', True)),
        },
    )
    config.full_clean()
    config.save()
    return config


def _send_budget_telegram_alert(user, message: str) -> bool:
    from telegram_bot.models import TelegramCredential
    from telegram_bot.services import TelegramService

    credential = TelegramCredential.objects.filter(user=user, ativo=True).first()
    if credential is None:
        return False

    chat_id = credential.get_chat_id()
    if not chat_id:
        return False

    try:
        service = TelegramService(credential.get_token())
        service.send_message(chat_id=chat_id, text=message)
        return True
    except Exception as exc:  # noqa: BLE001
        logger.warning('Falha ao enviar alerta de orçamento para Telegram: %s', exc)
        return False


def _build_budget_alert_message(event_type: str, status: dict, transaction: Transaction) -> str:
    if event_type == BudgetAlertEvent.EVENT_THRESHOLD_20:
        return (
            '⚠️ <b>Alerta de orçamento</b>\n\n'
            'Você consumiu 80% do limite mensal de saídas.\n'
            f'Limite: R$ {status["limite_mensal"]:.2f}\n'
            f'Gasto atual: R$ {status["total_saidas"]:.2f}\n'
            f'Restante: R$ {status["restante"]:.2f}'
        )
    if event_type == BudgetAlertEvent.EVENT_THRESHOLD_10:
        return (
            '🚨 <b>Alerta de orçamento</b>\n\n'
            'Faltam menos de 10% para seu limite mensal de saídas.\n'
            f'Limite: R$ {status["limite_mensal"]:.2f}\n'
            f'Gasto atual: R$ {status["total_saidas"]:.2f}\n'
            f'Restante: R$ {status["restante"]:.2f}'
        )
    if event_type == BudgetAlertEvent.EVENT_LIMIT_REACHED:
        return (
            '⛔ <b>Limite mensal atingido</b>\n\n'
            f'Limite: R$ {status["limite_mensal"]:.2f}\n'
            f'Gasto atual: R$ {status["total_saidas"]:.2f}\n'
            'A partir de agora, novas despesas disparam alerta.'
        )
    return (
        '❗ <b>Despesa acima do limite</b>\n\n'
        f'Nova despesa registrada: R$ {transaction.valor:.2f}\n'
        f'Descrição: {transaction.descricao[:80]}\n'
        f'Total gasto no mês: R$ {status["total_saidas"]:.2f}\n'
        f'Limite mensal: R$ {status["limite_mensal"]:.2f}'
    )


def evaluate_and_dispatch_budget_alerts(transaction: Transaction) -> None:
    from .selectors import get_budget_status

    if transaction.tipo != 'saida':
        return

    config = MonthlyBudgetConfig.objects.filter(usuario=transaction.usuario).first()
    if config is None or not config.alertas_ativos:
        return

    status = get_budget_status(transaction.usuario, reference_date=transaction.data)
    if status is None:
        return

    consumed = status['consumo_percentual']
    year = transaction.data.year
    month = transaction.data.month

    if consumed >= Decimal('80'):
        event_20, created = BudgetAlertEvent.objects.get_or_create(
            usuario=transaction.usuario,
            ano=year,
            mes=month,
            event_type=BudgetAlertEvent.EVENT_THRESHOLD_20,
            defaults={
                'mensagem': _build_budget_alert_message(
                    BudgetAlertEvent.EVENT_THRESHOLD_20,
                    status,
                    transaction,
                )
            },
        )
        if created:
            _send_budget_telegram_alert(transaction.usuario, event_20.mensagem)

    if consumed >= Decimal('90'):
        event_10, created = BudgetAlertEvent.objects.get_or_create(
            usuario=transaction.usuario,
            ano=year,
            mes=month,
            event_type=BudgetAlertEvent.EVENT_THRESHOLD_10,
            defaults={
                'mensagem': _build_budget_alert_message(
                    BudgetAlertEvent.EVENT_THRESHOLD_10,
                    status,
                    transaction,
                )
            },
        )
        if created:
            _send_budget_telegram_alert(transaction.usuario, event_10.mensagem)

    if consumed >= Decimal('100'):
        event_limit, created = BudgetAlertEvent.objects.get_or_create(
            usuario=transaction.usuario,
            ano=year,
            mes=month,
            event_type=BudgetAlertEvent.EVENT_LIMIT_REACHED,
            defaults={
                'mensagem': _build_budget_alert_message(
                    BudgetAlertEvent.EVENT_LIMIT_REACHED,
                    status,
                    transaction,
                )
            },
        )
        if created:
            _send_budget_telegram_alert(transaction.usuario, event_limit.mensagem)

    if consumed > Decimal('100'):
        post_limit_message = _build_budget_alert_message(
            BudgetAlertEvent.EVENT_POST_LIMIT_EXPENSE,
            status,
            transaction,
        )
        BudgetAlertEvent.objects.create(
            usuario=transaction.usuario,
            ano=year,
            mes=month,
            event_type=BudgetAlertEvent.EVENT_POST_LIMIT_EXPENSE,
            mensagem=post_limit_message,
        )
        _send_budget_telegram_alert(transaction.usuario, post_limit_message)


# ---------------------------------------------------------------------------
# Relatórios — helpers
# ---------------------------------------------------------------------------

def _format_brl(value) -> str:
    """Formata valor Decimal para padrão monetário brasileiro: R$ 1.234,56."""
    if value is None:
        return 'R$ 0,00'
    v = Decimal(str(value))
    sign = '-' if v < 0 else ''
    v = abs(v)
    int_part = int(v)
    dec_part = int(round((v - int_part) * 100))
    int_str = f'{int_part:,}'.replace(',', '.')
    return f'{sign}R$ {int_str},{dec_part:02d}'


# ---------------------------------------------------------------------------
# Relatórios — PDF
# ---------------------------------------------------------------------------

def generate_pdf_report(user, data_inicio, data_fim) -> bytes:
    """Gera relatório financeiro em PDF usando reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from .selectors import get_report_data

    data = get_report_data(user, data_inicio, data_fim)
    resumo = data['resumo']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # --- Título ---
    title_style = ParagraphStyle(
        'Title', parent=styles['Title'], fontSize=18, spaceAfter=6,
    )
    elements.append(Paragraph('FinTrack — Relatório Financeiro', title_style))
    periodo_str = (
        f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
    )
    elements.append(Paragraph(f'Período: {periodo_str}', styles['Normal']))

    import datetime

    elements.append(Paragraph(
        f'Gerado em: {datetime.date.today().strftime("%d/%m/%Y")}',
        styles['Normal'],
    ))
    elements.append(Spacer(1, 10 * mm))

    # --- Resumo ---
    green = colors.HexColor('#2e7d32')
    red = colors.HexColor('#c62828')
    blue = colors.HexColor('#1565c0')

    saldo_color = blue if resumo['saldo_liquido'] >= 0 else red

    resumo_data = [
        ['Indicador', 'Valor'],
        ['Total Entradas', _format_brl(resumo['total_entradas'])],
        ['Total Saídas', _format_brl(resumo['total_saidas'])],
        ['Saldo Líquido', _format_brl(resumo['saldo_liquido'])],
        ['Nº de Transações', str(resumo['num_transacoes'])],
    ]
    resumo_table = Table(resumo_data, colWidths=[200, 200])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TEXTCOLOR', (1, 1), (1, 1), green),
        ('TEXTCOLOR', (1, 2), (1, 2), red),
        ('TEXTCOLOR', (1, 3), (1, 3), saldo_color),
    ]))
    elements.append(Paragraph(
        '<b>Resumo Financeiro</b>', styles['Heading2'],
    ))
    elements.append(resumo_table)
    elements.append(Spacer(1, 8 * mm))

    # --- Distribuição por Categoria ---
    if data['por_categoria']:
        elements.append(Paragraph(
            '<b>Distribuição por Categoria (Saídas)</b>',
            styles['Heading2'],
        ))
        cat_header = ['Categoria', 'Valor (R$)', 'Percentual (%)']
        cat_rows = [cat_header]
        for item in data['por_categoria']:
            cat_rows.append([
                item['categoria'],
                _format_brl(item['total']),
                f'{item["percentual"]:.2f}%',
            ])
        cat_table = Table(cat_rows, colWidths=[180, 150, 100])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 8 * mm))

    # --- Transações detalhadas ---
    elements.append(Paragraph(
        '<b>Transações Detalhadas</b>', styles['Heading2'],
    ))
    tx_header = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Conta', 'Valor (R$)']
    tx_rows = [tx_header]

    # Agrupar por categoria
    by_cat = {}
    for t in data['transacoes']:
        cat_name = t.categoria.nome if t.categoria else 'Sem categoria'
        by_cat.setdefault(cat_name, []).append(t)

    for cat_name in sorted(by_cat.keys()):
        txns = by_cat[cat_name]
        subtotal = Decimal('0.00')
        for t in txns:
            tipo_display = 'Entrada' if t.tipo == 'entrada' else 'Saída'
            tx_rows.append([
                t.data.strftime('%d/%m/%Y'),
                tipo_display,
                t.descricao[:40],
                cat_name,
                t.conta.nome if t.conta else '—',
                _format_brl(t.valor),
            ])
            subtotal += t.valor
        tx_rows.append(['', '', '', f'Subtotal {cat_name}', '', _format_brl(subtotal)])

    tx_table = Table(tx_rows, colWidths=[65, 50, 130, 90, 70, 80])
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]
    # Highlight subtotal rows
    for i, row in enumerate(tx_rows):
        if row[3] and str(row[3]).startswith('Subtotal'):
            style_cmds.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#eeeeee')))
    tx_table.setStyle(TableStyle(style_cmds))
    elements.append(tx_table)
    elements.append(Spacer(1, 10 * mm))

    # --- Rodapé ---
    elements.append(Paragraph(
        f'Relatório gerado por FinTrack em {datetime.date.today().strftime("%d/%m/%Y")}',
        styles['Normal'],
    ))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Relatórios — Excel
# ---------------------------------------------------------------------------

def generate_excel_report(user, data_inicio, data_fim) -> bytes:
    """Gera relatório financeiro em Excel (XLSX) usando openpyxl."""
    import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .selectors import get_report_data

    data = get_report_data(user, data_inicio, data_fim)
    resumo = data['resumo']

    wb = Workbook()

    # ====== Aba Resumo ======
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='333333')
    bold_font = Font(bold=True)
    green_font = Font(bold=True, color='2E7D32')
    red_font = Font(bold=True, color='C62828')
    blue_font = Font(bold=True, color='1565C0')

    ws_resumo.merge_cells('A1:B1')
    ws_resumo['A1'] = 'Relatório Financeiro — FinTrack'
    ws_resumo['A1'].font = header_font
    ws_resumo['A1'].fill = header_fill

    periodo_str = (
        f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
    )
    ws_resumo['A3'] = 'Período:'
    ws_resumo['A3'].font = bold_font
    ws_resumo['B3'] = periodo_str

    kpis = [
        ('Total Entradas', resumo['total_entradas'], green_font),
        ('Total Saídas', resumo['total_saidas'], red_font),
        ('Saldo Líquido', resumo['saldo_liquido'],
         blue_font if resumo['saldo_liquido'] >= 0 else red_font),
        ('Nº Transações', resumo['num_transacoes'], bold_font),
        ('Maior Entrada', resumo['maior_entrada'], green_font),
        ('Maior Saída', resumo['maior_saida'], red_font),
    ]
    for i, (label, value, font) in enumerate(kpis, start=5):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'A{i}'].font = bold_font
        cell = ws_resumo[f'B{i}']
        if isinstance(value, Decimal):
            cell.value = float(value)
            cell.number_format = '#,##0.00'
        elif value is None:
            cell.value = '—'
        else:
            cell.value = value
        cell.font = font

    ws_resumo.column_dimensions['A'].width = 25
    ws_resumo.column_dimensions['B'].width = 25

    # ====== Aba Transações ======
    ws_tx = wb.create_sheet('Transações')
    tx_headers = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Conta', 'Valor']
    for col, h in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, t in enumerate(data['transacoes'], start=2):
        ws_tx.cell(row=row_idx, column=1, value=t.data.strftime('%d/%m/%Y'))
        ws_tx.cell(
            row=row_idx, column=2,
            value='Entrada' if t.tipo == 'entrada' else 'Saída',
        )
        ws_tx.cell(row=row_idx, column=3, value=t.descricao)
        ws_tx.cell(
            row=row_idx, column=4,
            value=t.categoria.nome if t.categoria else 'Sem categoria',
        )
        ws_tx.cell(
            row=row_idx, column=5,
            value=t.conta.nome if t.conta else '—',
        )
        valor_cell = ws_tx.cell(row=row_idx, column=6, value=float(t.valor))
        valor_cell.number_format = '#,##0.00'

    # Auto-filter
    if data['transacoes'].exists():
        last_row = ws_tx.max_row
        ws_tx.auto_filter.ref = f'A1:F{last_row}'

    for col in range(1, 7):
        ws_tx.column_dimensions[get_column_letter(col)].width = 18

    # ====== Aba Por Categoria ======
    ws_cat = wb.create_sheet('Por Categoria')
    cat_headers = ['Categoria', 'Total (R$)', 'Percentual (%)']
    for col, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(data['por_categoria'], start=2):
        ws_cat.cell(row=row_idx, column=1, value=item['categoria'])
        total_cell = ws_cat.cell(
            row=row_idx, column=2, value=float(item['total']),
        )
        total_cell.number_format = '#,##0.00'
        ws_cat.cell(row=row_idx, column=3, value=item['percentual'])

    ws_cat.column_dimensions['A'].width = 25
    ws_cat.column_dimensions['B'].width = 20
    ws_cat.column_dimensions['C'].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
